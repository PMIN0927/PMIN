from __future__ import annotations

import logging
import sys

from openpyxl import load_workbook

from .config import (
    FAILED_CSV_PATH,
    INPUT_XLSX_PATH,
    OUTPUT_CSV_PATH,
    OUTPUT_DIR,
    OUTPUT_XLSX_PATH,
    REVIEW_CSV_PATH,
    SHEET_PRIORITY,
    SUMMARY_TXT_PATH,
    load_settings,
)
from .coordinate_matcher import match_coordinates
from .export_excel import export_outputs
from .kakao_client import KakaoLocalClient


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s - %(message)s", handlers=[logging.StreamHandler(sys.stdout)])


def load_input_rows() -> tuple[list[dict], list[str], str]:
    if not INPUT_XLSX_PATH.exists():
        raise FileNotFoundError(f"입력 파일이 없습니다: {INPUT_XLSX_PATH}")
    wb = load_workbook(INPUT_XLSX_PATH, read_only=True, data_only=True)
    sheet = None
    for name in SHEET_PRIORITY:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows_iter)]
    rows = []
    for values in rows_iter:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers)) if headers[i]}
        if any(value not in [None, ""] for value in row.values()):
            rows.append(row)
    return rows, [header for header in headers if header], sheet.title


def main() -> None:
    setup_logging()
    logger = logging.getLogger(__name__)
    settings = load_settings()
    rows, columns, sheet_name = load_input_rows()
    logger.info("입력 시트 로드: %s / %s행", sheet_name, len(rows))
    client = KakaoLocalClient(settings["kakao_rest_api_key"])
    matched_rows = match_coordinates(rows, client)
    export_outputs(
        matched_rows,
        columns,
        {
            "output_dir": OUTPUT_DIR,
            "output_xlsx": OUTPUT_XLSX_PATH,
            "output_csv": OUTPUT_CSV_PATH,
            "failed_csv": FAILED_CSV_PATH,
            "review_csv": REVIEW_CSV_PATH,
            "summary_txt": SUMMARY_TXT_PATH,
        },
    )
    logger.info("완료: %s", OUTPUT_XLSX_PATH)


if __name__ == "__main__":
    main()
