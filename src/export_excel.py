from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import ADDED_COLUMNS, OUTPUT_CSV_PATH
from .text_utils import infer_role


def export_outputs(rows: list[dict], original_columns: list[str], paths: dict) -> None:
    paths["output_dir"].mkdir(parents=True, exist_ok=True)
    columns = _ordered_columns(original_columns)
    _write_csv(paths["output_csv"], rows, columns)
    _write_csv(paths["failed_csv"], [row for row in rows if row.get("coordinate_review_status") == "검색실패"], columns)
    _write_csv(paths["review_csv"], [row for row in rows if row.get("coordinate_review_status") in ["검수필요", "검색실패"]], columns)
    summary = build_summary(rows)
    paths["summary_txt"].write_text(summary_text(summary), encoding="utf-8")
    _write_xlsx(paths["output_xlsx"], rows, columns, summary)


def _ordered_columns(original_columns: list[str]) -> list[str]:
    columns = list(original_columns)
    for column in ADDED_COLUMNS:
        if column not in columns:
            columns.append(column)
    return columns


def _write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict], columns: list[str], summary: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("전체_좌표추가", rows, columns),
        ("확인완료", [r for r in rows if r.get("coordinate_review_status") == "확인완료"], columns),
        ("후보확인", [r for r in rows if r.get("coordinate_review_status") == "후보확인"], columns),
        ("검수필요", [r for r in rows if r.get("coordinate_review_status") == "검수필요"], columns),
        ("검색실패", [r for r in rows if r.get("coordinate_review_status") == "검색실패"], columns),
        ("식사", [r for r in rows if _role(r) == "식사"], columns),
        ("술", [r for r in rows if _role(r) == "술"], columns),
        ("카페", [r for r in rows if _role(r) == "카페"], columns),
        ("중간경유지", [r for r in rows if _role(r) == "중간경유지"], columns),
        ("요약", [{"항목": k, "값": v} for k, v in summary.items()], ["항목", "값"]),
    ]
    for name, sheet_rows, sheet_columns in sheets:
        ws = wb.create_sheet(name)
        ws.append(sheet_columns)
        for row in sheet_rows:
            ws.append([row.get(column, "") for column in sheet_columns])
        _style_sheet(ws)
    wb.save(path)


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="174A7C")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)
    ws.auto_filter.ref = ws.dimensions


def _role(row: dict) -> str:
    return row.get("장소역할") or row.get("place_role") or infer_role(row)


def build_summary(rows: list[dict]) -> dict:
    return {
        "전체 장소 수": len(rows),
        "좌표 확인완료 개수": len([r for r in rows if r.get("coordinate_review_status") == "확인완료"]),
        "후보확인 개수": len([r for r in rows if r.get("coordinate_review_status") == "후보확인"]),
        "검수필요 개수": len([r for r in rows if r.get("coordinate_review_status") == "검수필요"]),
        "검색실패 개수": len([r for r in rows if r.get("coordinate_review_status") == "검색실패"]),
        "식사 개수": len([r for r in rows if _role(r) == "식사"]),
        "술 개수": len([r for r in rows if _role(r) == "술"]),
        "카페 개수": len([r for r in rows if _role(r) == "카페"]),
        "중간경유지 개수": len([r for r in rows if _role(r) == "중간경유지"]),
        "address_search 개수": len([r for r in rows if r.get("coordinate_source") == "address_search"]),
        "keyword_search 개수": len([r for r in rows if r.get("coordinate_source") == "keyword_search"]),
        "failed 개수": len([r for r in rows if r.get("coordinate_source") == "failed"]),
    }


def summary_text(summary: dict) -> str:
    return "\n".join(f"{key}: {value}" for key, value in summary.items())
